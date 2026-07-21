import os
import io
import re
import html
import json
import shutil
import zipfile
import tempfile
import datetime as dt
import pandas as pd
import anthropic
import streamlit as st
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "단순경비율.xlsx")
ACCOUNT_PATH = os.path.join(BASE_DIR, "계정과목.xlsx")

PURCHASE_CATEGORIES = {"도매 및 상품중개업", "소매업", "음식점"}

EXCLUDED_ACCOUNTS = {
    "인적용역": {"감가상각비", "임차료", "차량유지비", "복리후생비", "운반비"},
    "운수": {"급료", "임차료", "복리후생비", "광고선전비", "운반비"},
    "운송": {"급료", "임차료", "복리후생비", "광고선전비", "운반비"},
}

# 종합소득세율 구간 (과세표준, 세율, 누진공제)
TAX_BRACKETS = [
    (14_000_000,   0.06,         0),
    (50_000_000,   0.15, 1_260_000),
    (88_000_000,   0.24, 5_760_000),
    (150_000_000,  0.35, 15_440_000),
    (300_000_000,  0.38, 19_940_000),
    (500_000_000,  0.40, 25_940_000),
    (1_000_000_000,0.42, 35_940_000),
    (float("inf"), 0.45, 65_940_000),
]


@st.cache_data
def load_data():
    df = pd.read_excel(EXCEL_PATH, header=None)
    df.columns = ["업종코드", "중분류", "소분류", "세분류", "세세분류",
                  "단순일반율", "단순자가율", "기준일반율", "기준자가율", "기준및적용범위"]
    df["업종코드"] = df["업종코드"].astype(str).str.strip()
    return df


@st.cache_data
def load_accounts():
    df = pd.read_excel(ACCOUNT_PATH, header=None)
    accounts = {"매출원가": [], "판관비": []}
    current = None
    for _, row in df.iterrows():
        if pd.notna(row[0]):
            current = str(row[0]).strip()
        if current and pd.notna(row[1]):
            accounts[current].append(str(row[1]).strip())
    return accounts


def find_business(df, code):
    row = df[df["업종코드"] == code]
    if row.empty:
        return None
    return row.iloc[0]


def has_product_purchase(중분류):
    return any(cat in str(중분류) for cat in PURCHASE_CATEGORIES)


def get_excluded_accounts(중분류):
    excluded = set()
    for key, accs in EXCLUDED_ACCOUNTS.items():
        if key in str(중분류):
            excluded |= accs
    return excluded


def get_expense_distribution(business_info, accounts, use_purchase, 적용경비율=None):
    api_key = st.secrets.get("ANTHROPIC_API_KEY") or os.getenv("ANTHROPIC_API_KEY")
    client = anthropic.Anthropic(api_key=api_key)

    excluded = get_excluded_accounts(business_info["중분류"])
    account_list = [a for a in accounts["판관비"] if a not in excluded]

    if use_purchase:
        all_accounts = ["당기상품매입액"] + account_list
    else:
        all_accounts = account_list

    expense_rate = 적용경비율 if 적용경비율 is not None else float(business_info["단순일반율"])

    prompt = f"""당신은 세무 전문가입니다. 아래 업종의 경비율을 계정과목별로 배분해주세요.

업종 정보:
- 업종코드: {business_info['업종코드']}
- 중분류: {business_info['중분류']}
- 소분류: {business_info['소분류']}
- 세분류: {business_info['세분류']}
- 세세분류: {business_info['세세분류']}
- 단순경비율(일반): {expense_rate}%
- 업종 설명: {business_info['기준및적용범위']}

배분할 계정과목:
{json.dumps(all_accounts, ensure_ascii=False)}

조건:
1. 위 계정과목들의 비율 합계가 정확히 {expense_rate}%가 되어야 합니다
2. 해당 업종의 특성에 맞게 현실적으로 배분해주세요
3. 해당 업종에서 발생하지 않는 비용은 0%로 설정하세요
4. 소수점 첫째 자리까지만 표시하세요

반드시 아래 JSON 형식으로만 응답하세요 (설명 없이):
{{
  "계정과목명": 비율숫자,
  ...
}}"""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}]
    )

    response_text = message.content[0].text.strip()
    if response_text.startswith("```"):
        response_text = response_text.split("```")[1]
        if response_text.startswith("json"):
            response_text = response_text[4:]

    return json.loads(response_text.strip())


def calc_income_tax(과세표준):
    """과세표준으로 산출세액 계산"""
    for limit, rate, deduction in TAX_BRACKETS:
        if 과세표준 <= limit:
            return round(과세표준 * rate - deduction)
    return 0


# ── 페이지 설정 ──────────────────────────────────────
st.set_page_config(page_title="단순경비율 계산기", page_icon="📊", layout="centered")

# ── 비밀번호 인증 ──────────────────────────────────────
def check_password():
    correct_pw = st.secrets.get("APP_PASSWORD") or os.getenv("APP_PASSWORD", "tax1234")

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.markdown("---")
    st.markdown("<h2 style='text-align:center'>🔐 경비율 계산기</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align:center; color:gray'>이용 문의: 이재원 세무사</p>", unsafe_allow_html=True)
    st.markdown("---")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pw = st.text_input("비밀번호를 입력하세요", type="password", placeholder="비밀번호")
        if st.button("입력", use_container_width=True, type="primary"):
            if pw == correct_pw:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")
    return False

if not check_password():
    st.stop()

# ══════════════════════════════════════════════════════
# 카드사 변환 관련 함수
# ══════════════════════════════════════════════════════
TEMPLATE_PATH = os.path.join(BASE_DIR, "신용카드매입자료_업로드_기본서식.xlsx")

ACCOUNT_CODE_MAP = {
    "복리후생비": "811", "여비교통비": "812", "접대비": "813",
    "통신비": "814", "수도광열비": "815", "전력비": "816",
    "세금과공과금": "817", "임차료": "819", "수선비": "820",
    "보험료": "821", "차량유지비": "822", "운반비": "824",
    "교육훈련비": "825", "도서인쇄비": "826", "회의비": "827",
    "포장비": "828", "사무용품비": "829", "소모품비": "830",
    "수수료비용": "831", "보관료": "832", "광고선전비": "833",
    "건물관리비": "837", "미분류": "",
}

DEFAULT_RULES = {
    "keyword": [
        {"match": ["주유소","SK에너지","GS칼텍스","현대오일뱅크","S-OIL"], "account": "차량유지비", "vat": "공제"},
        {"match": ["정비","카센터","타이어","세차"], "account": "차량유지비", "vat": "공제"},
        {"match": ["통행료","하이패스","주차","한국도로공사"], "account": "차량유지비", "vat": "공제"},
        {"match": ["충전","일렉링크","전기차충전"], "account": "차량유지비", "vat": "공제"},
        {"match": ["GS25","CU","세븐일레븐","이마트24","미니스톱","씨유"], "account": "복리후생비", "vat": "공제"},
        {"match": ["이마트","홈플러스","롯데마트"], "account": "복리후생비", "vat": "공제"},
        {"match": ["맥도날드","버거킹","롯데리아","KFC","한솥도시락"], "account": "복리후생비", "vat": "공제"},
        {"match": ["스타벅스","커피","카페","빽다방","투썸"], "account": "접대비", "vat": "불공제"},
        {"match": ["SK텔레콤","KT","LG유플러스","SKT","LGU+"], "account": "통신비", "vat": "공제"},
        {"match": ["호텔","모텔","펜션","아난티"], "account": "여비교통비", "vat": "공제"},
        {"match": ["택시","카카오T"], "account": "여비교통비", "vat": "공제"},
        {"match": ["병원","의원","한의원","치과"], "account": "복리후생비", "vat": "불공제"},
        {"match": ["약국"], "account": "복리후생비", "vat": "불공제"},
        {"match": ["우체국"], "account": "통신비", "vat": "불공제"},
    ],
    "industry": [
        {"match": ["주유소"], "account": "차량유지비", "vat": "공제"},
        {"match": ["편의점"], "account": "복리후생비", "vat": "공제"},
        {"match": ["음식점","한식","중식","일식","양식","부페"], "account": "접대비", "vat": "불공제"},
        {"match": ["커피전문점"], "account": "접대비", "vat": "불공제"},
        {"match": ["숙박업","호텔"], "account": "여비교통비", "vat": "공제"},
        {"match": ["통신사"], "account": "통신비", "vat": "공제"},
        {"match": ["통행료"], "account": "차량유지비", "vat": "공제"},
    ]
}


def parse_filename_card(filename):
    # 형식: 상호명_사업자번호_카드사_카드번호_직원유무_차량유무.xlsx
    # 카드사 생략 형식: 상호명_사업자번호_카드번호_직원유무_차량유무.xlsx
    base = os.path.splitext(filename)[0]
    parts = base.split("_")
    result = {"업체명": "", "사업자번호": "", "신용카드사명": "", "신용카드번호": "", "직원유무": "", "차량유무": ""}
    if len(parts) >= 1: result["업체명"] = parts[0]
    if len(parts) >= 2: result["사업자번호"] = parts[1]
    if len(parts) >= 3:
        # parts[2]가 카드번호 패턴(숫자4개-...)이면 카드사 생략으로 판단
        if parts[2].count("-") >= 3 and any(c.isdigit() for c in parts[2]):
            result["신용카드번호"] = parts[2]
            if len(parts) >= 4: result["직원유무"] = parts[3]
            if len(parts) >= 5: result["차량유무"] = parts[4]
        else:
            result["신용카드사명"] = parts[2]
            if len(parts) >= 4: result["신용카드번호"] = parts[3]
            if len(parts) >= 5: result["직원유무"] = parts[4]
            if len(parts) >= 6: result["차량유무"] = parts[5]
    return result


def classify_transaction(vendor, industry, rules):
    vendor = str(vendor) if vendor else ""
    industry = str(industry) if industry else ""
    for rule in rules["keyword"]:
        for kw in rule["match"]:
            if kw in vendor:
                return rule["account"], rule["vat"], "키워드"
    if industry and industry not in ("nan", ""):
        for rule in rules["industry"]:
            for kw in rule["match"]:
                if kw in industry:
                    return rule["account"], rule["vat"], "업종"
    return "미분류", "공제", "미분류"


def process_card_data(vendor, date, total, bizno, upjong, card_company, card_number):
    supply = (total / 1.1).astype(int)
    vat = total - supply
    accounts, vat_deds, methods = [], [], []
    for v, u in zip(vendor, upjong):
        acc, vd, m = classify_transaction(v, u, DEFAULT_RULES)
        accounts.append(acc); vat_deds.append(vd); methods.append(m)
    vat_types = [57 if v == "공제" else "" for v in vat_deds]
    account_codes = [ACCOUNT_CODE_MAP.get(a, "") for a in accounts]
    result = pd.DataFrame({
        "카드종류": 1, "신용카드사명": card_company, "신용카드번호": card_number,
        "승인일자": date.dt.strftime("%Y%m%d"), "사업자등록번호": bizno,
        "거래처명": vendor.str.slice(0, 15), "거래처유형": "",
        "공급가액": supply, "세액": vat, "봉사료": 0, "합계금액": total,
        "부가세공제여부": vat_deds, "부가세유형": vat_types,
        "계정과목": account_codes, "품목(적요)": upjong.str.slice(0, 30) if hasattr(upjong, "str") else "",
    })
    stats = pd.DataFrame({
        "가맹점명": vendor, "업종": upjong, "계정과목": accounts,
        "부가세": vat_deds, "분류방법": methods, "금액": total
    })
    return result, stats


def parse_samsung_card(file_bytes, card_company, card_number):
    # ── 신형 삼성카드: header=0, 매출금액(원) 컬럼 ──
    df_check0 = pd.read_excel(io.BytesIO(file_bytes), header=0, nrows=0)
    cols0 = [str(c).strip() for c in df_check0.columns]
    if '매출금액(원)' in cols0 and '매출일자' in cols0 and '가맹점명' in cols0:
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
        df.columns = [str(c).strip() for c in df.columns]
        date   = pd.to_datetime(df['매출일자'].astype(str), format='%Y%m%d', errors='coerce')
        vendor = df['가맹점명'].astype(str).str.strip()
        total  = pd.to_numeric(df['매출금액(원)'], errors='coerce').fillna(0).astype(int)
        bizno  = df['사업자등록번호'].astype(str).str.replace('-', '').str[:10] if '사업자등록번호' in df.columns else pd.Series([''] * len(df))
        upjong = pd.Series([''] * len(df))
        mask = date.notna() & (total > 0)
        return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                                 total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                                 upjong[mask].reset_index(drop=True), card_company, card_number)

    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None)

    # 헤더 행 탐색
    header_row = None
    for i in range(min(100, len(df_raw))):
        row_join = " ".join("" if x is None or (isinstance(x, float) and pd.isna(x)) else str(x) for x in df_raw.iloc[i].tolist())
        if ("이용일" in row_join or "승인일자" in row_join) and ("가맹점명" in row_join or "가맹점" in row_join) and "금액" in row_join:
            header_row = i; break
    if header_row is None: header_row = 19

    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    def pick(cands):
        for c in cands:
            if c in df.columns: return c
        return None

    col_v = pick(["가맹점명", "가맹점", "상호"])
    col_d = pick(["이용일", "승인일자", "거래일"])
    col_t = pick(["이용금액(원)", "이용금액", "금액"])
    col_b = pick(["사업자번호", "사업자등록번호"])
    col_u = pick(["업종", "업태"])

    # 컬럼명 감지 실패 시 위치(index) 기반으로 폴백
    if col_v is None and len(df.columns) >= 7:
        df.columns = list(df.columns)
        # 삼성카드 표준 12컬럼: 상품 매출 성명 카드구분 카드번호 이용일 가맹점명 이용금액(원) 개월수 승인번호 사업자번호 업종
        col_map = {0:"상품", 1:"매출", 2:"성명", 3:"카드구분", 4:"카드번호",
                   5:"이용일", 6:"가맹점명", 7:"이용금액(원)", 8:"개월수",
                   9:"승인번호", 10:"사업자번호", 11:"업종"}
        df.columns = [col_map.get(i, f"col{i}") for i in range(len(df.columns))]
        col_v = "가맹점명"; col_d = "이용일"; col_t = "이용금액(원)"
        col_b = "사업자번호"; col_u = "업종"

    if col_v is None or col_d is None or col_t is None:
        raise ValueError(f"삼성카드 컬럼 감지 실패. 발견된 컬럼: {list(df.columns)}")

    vendor = df[col_v].astype(str).str.replace(r"_x000D_", "", regex=True).str.strip()
    date = pd.to_datetime(df[col_d], errors="coerce")
    total = pd.to_numeric(df[col_t], errors="coerce").fillna(0).astype(int)
    bizno = df[col_b].apply(lambda x: re.sub(r"\D", "", str(x))[:10]) if col_b else pd.Series([""] * len(df))
    upjong = df[col_u].astype(str) if col_u else pd.Series([""] * len(df))

    mask = date.notna() & (total != 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_hana_card(file_bytes, card_company, card_number):
    """하나카드 파싱 - 접수일자/원화사용금액 형식, 승인일자/승인금액 형식, 구형 자동 감지"""

    # ── 종합소득세 형식: header=0, 접수일자/가맹점명/원화사용금액 ──
    df_h0 = pd.read_excel(io.BytesIO(file_bytes), header=0, nrows=0)
    cols_h0 = set(str(c).strip() for c in df_h0.columns)
    if {'접수일자', '가맹점명', '원화사용금액'}.issubset(cols_h0):
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
        df.columns = [str(c).strip() for c in df.columns]
        date   = pd.to_datetime(df['접수일자'], errors='coerce')
        vendor = df['가맹점명'].astype(str).str.strip()
        total  = pd.to_numeric(df['원화사용금액'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0).astype(int)
        bizno  = df['가맹점사업자번호'].astype(str).str.replace('-', '').str[:10] if '가맹점사업자번호' in df.columns else pd.Series([''] * len(df))
        upjong = df['업종명'].astype(str) if '업종명' in df.columns else pd.Series([''] * len(df))
        mask = date.notna() & (total > 0)
        return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                                 total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                                 upjong[mask].reset_index(drop=True), card_company, card_number)

    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)

    # 신형 형식: '승인일자' + '가맹점명' + '승인금액' 헤더가 있는 경우
    # 데이터가 날짜 패턴(YYYY-MM-DD)으로 시작하는 행만 추출
    HEADER_MARKER = {"승인일자", "가맹점명", "승인금액"}
    is_new_format = False
    for _, row in df_raw.iterrows():
        vals = set(str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ["", "nan"])
        if HEADER_MARKER.issubset(vals):
            is_new_format = True
            break

    if is_new_format:
        # 헤더 행에서 컬럼 위치(인덱스) 동적 추출
        col_idx = {}
        for _, row in df_raw.iterrows():
            vals_raw = [str(v).strip() if pd.notna(v) else "" for v in row]
            if "승인일자" in vals_raw and "가맹점명" in vals_raw:
                for ci, v in enumerate(vals_raw):
                    col_idx[v] = ci
                break
        ci_date   = col_idx.get("승인일자", 0)
        ci_vendor = col_idx.get("가맹점명", 6)
        ci_bizno  = col_idx.get("사업자번호", 9)
        ci_amount = col_idx.get("승인금액", 15)
        ci_cancel = col_idx.get("취소여부", 16)

        # 날짜 패턴 행만 수집, 취소 행 제외
        DATE_PAT = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        vendors, dates, totals, biznos, upjongnums = [], [], [], [], []
        for _, row in df_raw.iterrows():
            vals = [str(v).strip() if pd.notna(v) else "" for v in row]
            if not DATE_PAT.match(vals[ci_date] if ci_date < len(vals) else ""):
                continue
            if ci_cancel < len(vals) and vals[ci_cancel] == "취소":
                continue
            try:
                amount = int(str(vals[ci_amount] if ci_amount < len(vals) else "0").replace(",", ""))
            except ValueError:
                continue
            if amount <= 0:
                continue
            dates.append(vals[ci_date])
            vendors.append(vals[ci_vendor] if ci_vendor < len(vals) else "")
            biznos.append(re.sub(r"\D", "", vals[ci_bizno] if ci_bizno < len(vals) else "")[:10])
            totals.append(amount)
            upjongnums.append("")
        if not vendors:
            return pd.DataFrame(), pd.DataFrame()
        date_s  = pd.to_datetime(pd.Series(dates), errors="coerce")
        vendor_s = pd.Series(vendors)
        total_s  = pd.Series(totals, dtype=int)
        bizno_s  = pd.Series(biznos)
        upjong_s = pd.Series(upjongnums)
        mask = date_s.notna() & (total_s != 0)
        return process_card_data(vendor_s[mask].reset_index(drop=True), date_s[mask].reset_index(drop=True),
                                 total_s[mask].reset_index(drop=True), bizno_s[mask].reset_index(drop=True),
                                 upjong_s[mask].reset_index(drop=True), card_company, card_number)

    # 구형 형식: header=12, 매출일자/원화사용금액 컬럼
    df = pd.read_excel(io.BytesIO(file_bytes), header=12)
    df = df[df['취소일자'].isna() | (df['취소일자'] == '취소일자')]
    vendor = df['가맹점명'].astype(str).str.strip()
    date = pd.to_datetime(df['매출일자'], errors="coerce")
    total = pd.to_numeric(df['원화\n사용금액'], errors="coerce").fillna(0).astype(int)
    bizno = df['가맹점\n사업자번호'].astype(str).str.replace("-", "").str[:10]
    upjong = df['업종명'].astype(str) if '업종명' in df.columns else pd.Series([""] * len(df))
    mask = date.notna() & (total != 0)
    return process_card_data(vendor[mask], date[mask], total[mask], bizno[mask], upjong[mask], card_company, card_number)


def parse_shinhan_card(file_bytes, card_company, card_number):
    """신한카드 파싱 - 셀 값 기반 헤더 행 자동 탐색, 컬럼명 기반 추출"""
    HEADER_KEYWORDS = ["가맹점명", "매출금액", "이용금액", "거래금액", "이용일", "거래일"]

    # header=None 으로 전체 읽어서 셀 값에 키워드가 2개 이상인 행을 헤더로 감지
    header_idx = 4  # 기본값
    try:
        df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
        for i, row in df_raw.iterrows():
            vals = " ".join(str(v) for v in row if pd.notna(v))
            matched = sum(1 for k in HEADER_KEYWORDS if k in vals)
            if matched >= 2:
                header_idx = i
                break
    except Exception:
        pass

    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    cols = [str(c).strip() for c in df.columns]
    ncols = len(df.columns)

    def find_col(candidates):
        for c in candidates:
            for i, col in enumerate(cols):
                if c in col:
                    return i
        return None

    def safe_col(idx, default_idx):
        i = idx if idx is not None else default_idx
        if i is not None and 0 <= i < ncols:
            return df.iloc[:, i]
        return pd.Series([""] * len(df))

    date_idx   = find_col(["이용일", "거래일"])
    vendor_idx = find_col(["가맹점명", "가맹점"])
    amount_idx = find_col(["매출금액", "이용금액", "거래금액", "승인금액"])
    bizno_idx  = find_col(["사업자등록번호", "사업자번호"])
    upjong_idx = find_col(["상품구분", "상품유형", "업종"])

    date   = pd.to_datetime(safe_col(date_idx, 0), errors="coerce")
    vendor = safe_col(vendor_idx, 4).astype(str).str.strip()
    total  = pd.to_numeric(safe_col(amount_idx, 6), errors="coerce").fillna(0).astype(int)
    bizno  = safe_col(bizno_idx, 5).astype(str).str.replace("-", "").str[:10]
    upjong = safe_col(upjong_idx, 3).astype(str)

    mask = date.notna() & (total != 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_kb_card(file_bytes, card_company, card_number):
    """국민카드 파싱 - 이용일/매출금액/가맹점명/사업자번호, 취소여부 필터"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    header_idx = 13
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '이용일' in vals and '가맹점명' in vals and '매출금액' in vals:
            header_idx = i
            break
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    if '취소여부' in df.columns:
        df = df[df['취소여부'] == '정상'].reset_index(drop=True)
    date   = pd.to_datetime(df['이용일'], errors='coerce')
    vendor = df['가맹점명'].astype(str).str.strip()
    total  = pd.to_numeric(df['매출금액'], errors='coerce').fillna(0).astype(int)
    bizno  = df['사업자번호'].astype(str).str.replace('-', '').str[:10] if '사업자번호' in df.columns else pd.Series([''] * len(df))
    upjong = pd.Series([''] * len(df))
    mask = date.notna() & (total > 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_lotte_card(file_bytes, card_company, card_number):
    """롯데카드 파싱 - 매출일자/가맹점명/매출금액/사업자번호, 음수(취소)·합계행 제외"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    header_idx = 5
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '매출일자' in vals and '가맹점명' in vals and '가맹점번호' in vals:
            header_idx = i
            break
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    date   = pd.to_datetime(df['매출일자'], errors='coerce')
    vendor = df['가맹점명'].astype(str).str.strip()
    total  = pd.to_numeric(df['매출금액'], errors='coerce').fillna(0).astype(int)
    bizno  = df['사업자번호'].astype(str).str.replace('-', '').str[:10] if '사업자번호' in df.columns else pd.Series([''] * len(df))
    upjong = pd.Series([''] * len(df))
    mask = date.notna() & (total > 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_kakao_card(file_bytes, card_company, card_number):
    """카카오뱅크 카드이용내역 파싱 - 거래일시/가맹점명/매출금액/사업자등록번호, 취소여부 필터"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    header_idx = 16
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '거래일시' in vals and '가맹점명' in vals and '취소여부' in vals:
            header_idx = i
            break
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    df.columns = [str(c).strip().replace('\n', '') for c in df.columns]
    if '취소여부' in df.columns:
        df = df[df['취소여부'] == '정상'].reset_index(drop=True)
    date   = pd.to_datetime(df['거래일시'].astype(str).str[:10].str.replace('.', '-', regex=False), errors='coerce')
    vendor = df['가맹점명'].astype(str).str.strip()
    total  = pd.to_numeric(df['매출금액'].astype(str).str.replace(',', '', regex=False), errors='coerce').fillna(0).astype(int)
    bizno  = df['사업자등록번호'].astype(str).str.replace('-', '').str[:10] if '사업자등록번호' in df.columns else pd.Series([''] * len(df))
    upjong = pd.Series([''] * len(df))
    mask = date.notna() & (total > 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_ibk_bc_card(file_bytes, card_company, card_number):
    """IBK기업은행 BC카드 파싱 - 접수일자/가맹점명/이용금액/가맹점사업자번호, 헤더행 동적 감지"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    header_idx = 8
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '접수일자' in vals and '가맹점명' in vals and '이용금액' in vals:
            header_idx = i
            break
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    date_col = '접수일자' if '접수일자' in df.columns else '매출일자'
    date   = pd.to_datetime(df[date_col].astype(str), format='%Y%m%d', errors='coerce')
    vendor = df['가맹점명'].astype(str).str.strip()
    total  = pd.to_numeric(df['이용금액'], errors='coerce').fillna(0).astype(int)
    bizno  = df['가맹점사업자번호'].astype(str).str.replace('-', '').str[:10] if '가맹점사업자번호' in df.columns else pd.Series([''] * len(df))
    upjong = pd.Series([''] * len(df))
    mask = date.notna() & (total > 0)  # 음수(취소) 제외
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_hyundai_card(file_bytes, card_company, card_number):
    """현대카드 파싱 - YYYY.MM.DD 형식 및 Excel 시리얼 날짜, 이용금액 '원' 처리 포함"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    header_idx = 8
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '이용일' in vals and '가맹점명' in vals:
            header_idx = i
            break
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    if '상태' in df.columns:
        df = df[df['상태'] == '정상'].reset_index(drop=True)
    # 금액 컬럼: '이용 금액'(공백 있음), '이용금액'(공백 없음), '승인 금액', '승인금액' 순으로 탐색
    amount_col = next((c for c in ['이용 금액', '이용금액', '승인 금액', '승인금액'] if c in df.columns), None)
    if amount_col is None:
        raise ValueError(f"현대카드: 금액 컬럼을 찾을 수 없습니다. 컬럼: {list(df.columns)}")
    # 날짜: YYYY.MM.DD 형식 시도 후 대부분 NaT이면 Excel 시리얼 숫자로 재처리
    date = pd.to_datetime(df['이용일'].astype(str).str.replace('.', '-', regex=False), errors='coerce')
    if date.isna().mean() > 0.5:
        serial = pd.to_numeric(df['이용일'], errors='coerce')
        date = serial.apply(lambda x: pd.Timestamp('1899-12-30') + pd.Timedelta(days=x) if pd.notna(x) else pd.NaT)
    vendor = df['가맹점명'].astype(str).str.strip()
    # 금액: 쉼표·'원' 제거
    total  = pd.to_numeric(df[amount_col].astype(str).str.replace(',', '', regex=False).str.replace('원', '', regex=False).str.strip(), errors='coerce').fillna(0).astype(int)
    bizno_col = next((c for c in ['사업자등록번호', '사업자번호'] if c in df.columns), None)
    bizno  = df[bizno_col].astype(str).str.replace('-', '').str[:10] if bizno_col else pd.Series([''] * len(df))
    upjong = pd.Series([''] * len(df))
    mask = date.notna() & (total > 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_nh_card(file_bytes, card_company, card_number):
    """NH농협카드 파싱 - header=3, 매출일자/매출금액/사업자번호/가맹점명"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    # 헤더 행 탐색: '매출일자'+'가맹점명' 포함 행
    header_idx = 3
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '매출일자' in vals and '가맹점명' in vals:
            header_idx = i
            break
    df = pd.read_excel(io.BytesIO(file_bytes), header=header_idx)
    df.columns = [str(c).strip() for c in df.columns]
    date   = pd.to_datetime(df['매출일자'], errors='coerce')
    vendor = df['가맹점명'].astype(str).str.strip()
    total  = pd.to_numeric(df['매출금액'], errors='coerce').fillna(0).astype(int)
    bizno  = df['사업자번호'].astype(str).str.replace('-', '').str[:10] if '사업자번호' in df.columns else pd.Series([''] * len(df))
    upjong = pd.Series([''] * len(df))
    mask = date.notna() & (total > 0)
    return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                             total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                             upjong[mask].reset_index(drop=True), card_company, card_number)


def parse_bc_card(file_bytes, card_company, card_number):
    """비씨카드 파싱 - IBK기업은행형/신형(header=0)/구형 자동 감지"""

    # ── IBK기업은행 형식 감지: '접수일자' + '가맹점명' + '이용금액' ──
    df_raw_check = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
    for _i, _row in df_raw_check.iterrows():
        _vals = [str(v).strip() for v in _row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
        if '접수일자' in _vals and '가맹점명' in _vals and '이용금액' in _vals:
            return parse_ibk_bc_card(file_bytes, card_company, card_number)

    # ── 신형 형식 감지: 첫 행이 '매출일자' + '가맹점명' + '매출금액' 포함 ──
    df_check = pd.read_excel(io.BytesIO(file_bytes), header=0, nrows=0)
    cols0 = set(str(c).strip() for c in df_check.columns)
    if {"매출일자", "가맹점명", "매출금액"}.issubset(cols0):
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
        df.columns = [str(c).strip() for c in df.columns]
        # 소계/합계 행 제외: 고객사명 없는 행
        if "고객사명" in df.columns:
            df = df[df["고객사명"].notna()].reset_index(drop=True)
        # 날짜: "2026.01.01" 또는 일반 날짜 형식
        date   = pd.to_datetime(df["매출일자"].astype(str).str.replace('.', '-', regex=False), errors="coerce")
        vendor = df["가맹점명"].astype(str).str.strip()
        # 금액: "42,100 원" 형식 처리 (쉼표·'원' 제거)
        total  = pd.to_numeric(df["매출금액"].astype(str).str.replace(',', '').str.replace('원', '').str.strip(), errors="coerce").fillna(0).astype(int)
        bizno_col = next((c for c in ["사업자등록번호", "사업자번호"] if c in df.columns), None)
        bizno  = df[bizno_col].astype(str).str.replace("-", "").str[:10] if bizno_col else pd.Series([""] * len(df))
        upjong = pd.Series([""] * len(df))
        mask = date.notna() & (total > 0)
        return process_card_data(vendor[mask].reset_index(drop=True), date[mask].reset_index(drop=True),
                                 total[mask].reset_index(drop=True), bizno[mask].reset_index(drop=True),
                                 upjong[mask].reset_index(drop=True), card_company, card_number)

    # ── 구형 형식: YYYY/MM/DD 날짜 패턴으로 데이터행 식별 (openpyxl) ──
    from openpyxl import load_workbook as _load_wb
    wb = _load_wb(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx = 9
    for i, row in enumerate(all_rows):
        non_none = [c for c in row if c is not None]
        if non_none and re.match(r'\d{4}/\d{2}/\d{2}', str(non_none[0])):
            for j in range(i - 1, -1, -1):
                if any(c is not None for c in all_rows[j]):
                    header_idx = j
                    break
            break

    vendors, dates, totals, biznos = [], [], [], []
    for row in all_rows[header_idx + 1:]:
        date_val = row[2] if len(row) > 2 else None
        if date_val is None: continue
        date_str = str(date_val).strip()
        if not re.match(r'\d{4}/\d{2}/\d{2}', date_str): continue
        vendor = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ''
        amount_raw = row[19] if len(row) > 19 else None
        try:
            amount = int(str(amount_raw).replace(',', '').strip())
        except (ValueError, TypeError):
            amount = 0
        if amount <= 0: continue
        bizno_raw = str(row[12]).replace('-', '') if len(row) > 12 and row[12] is not None else ''
        biznos.append(re.sub(r'\D', '', bizno_raw)[:10])
        vendors.append(vendor); dates.append(date_str); totals.append(amount)

    if not vendors:
        return pd.DataFrame(), pd.DataFrame()

    vendor_s = pd.Series(vendors)
    date_s = pd.to_datetime(pd.Series(dates), format='%Y/%m/%d', errors='coerce')
    total_s = pd.Series(totals, dtype=int)
    bizno_s = pd.Series(biznos)
    upjong_s = pd.Series([""] * len(vendors))
    return process_card_data(vendor_s, date_s, total_s, bizno_s, upjong_s, card_company, card_number)


def fix_html_entities(xlsx_path):
    tmp = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z: z.extractall(tmp)
        for sub in ['xl/worksheets', 'xl']:
            d = os.path.join(tmp, sub)
            if not os.path.exists(d): continue
            for fn in os.listdir(d):
                if fn.endswith('.xml'):
                    fp = os.path.join(d, fn)
                    content = open(fp, encoding='utf-8').read()
                    open(fp, 'w', encoding='utf-8').write(html.unescape(content))
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for root, _, files in os.walk(tmp):
                for f in files: z.write(os.path.join(root, f), os.path.relpath(os.path.join(root, f), tmp))
    finally:
        shutil.rmtree(tmp)


def classify_unclassified_with_claude(vendor_names):
    """미분류 가맹점명을 Claude API로 일괄 분류 → {가맹점명: 계정과목명} 반환"""
    api_key = st.secrets.get("ANTHROPIC_API_KEY") or os.getenv("ANTHROPIC_API_KEY")
    client = anthropic.Anthropic(api_key=api_key)

    unique_vendors = list(dict.fromkeys(vendor_names))  # 순서 유지 중복 제거
    if not unique_vendors:
        return {}

    valid_accounts = [name for name in ACCOUNT_CODE_MAP if name != "미분류"]
    account_list_str = "\n".join(f"- {name}" for name in valid_accounts)

    prompt = f"""다음 신용카드 가맹점명을 아래 계정과목 중 가장 적합한 것으로 분류하세요.
업무용 법인카드/사업용 카드 이용 내역이며, 경비 처리 목적입니다.

[계정과목 목록]
{account_list_str}

[분류 기준]
- 식당, 음식점, 카페, 커피, 주점, 치킨, 피자, 횟집 → 접대비
- 편의점(GS25, CU, 세븐일레븐, 이마트24), 마트, 슈퍼, 약국, 병원 → 복리후생비
- 주유소, 충전소, 정비소, 세차, 주차장, 하이패스 → 차량유지비
- KT, SKT, LG유플러스, SK브로드밴드, 인터넷 → 통신비
- 항공(대한항공, 아시아나, 제주항공), 기차(KTX, 코레일), 버스, 택시, 숙박(호텔, 모텔, 여관) → 여비교통비
- 보험사(삼성생명, 한화생명, 흥국화재 등) → 보험료
- 학원, 교육기관, 훈련센터 → 교육훈련비
- 도서, 신문, 문구점, 사무용품 → 도서인쇄비 또는 사무용품비
- 임대업체, 월세, 건물 관리 → 임차료
- 분류 어려운 경우 → 소모품비

[가맹점명 목록 (JSON 배열)]
{json.dumps(unique_vendors, ensure_ascii=False)}

JSON 형식으로만 응답하세요 (설명 없이, 코드블록 없이):
{{"가맹점명": "계정과목명", ...}}"""

    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}]
    )
    text = msg.content[0].text.strip()
    # 코드블록 제거
    if "```" in text:
        parts = text.split("```")
        text = parts[1] if len(parts) > 1 else parts[0]
        if text.startswith("json"):
            text = text[4:]
    result = json.loads(text.strip())
    # 유효한 계정과목만 유지
    return {k: v for k, v in result.items() if v in ACCOUNT_CODE_MAP and v != "미분류"}


def write_to_template(rows_df, company_name, business_number):
    tmp_path = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(TEMPLATE_PATH, tmp_path)
    wb = load_workbook(tmp_path)
    ws = wb.active
    ws['A4'] = company_name; ws['C4'] = business_number
    for r in range(10, 10 + len(rows_df) + 10):
        for c in range(1, 16): ws.cell(r, c).value = None
    for idx, row in rows_df.iterrows():
        r = 10 + list(rows_df.index).index(idx)
        ws.cell(r, 1).value = int(row['카드종류']); ws.cell(r, 2).value = row['신용카드사명']
        ws.cell(r, 3).value = row['신용카드번호']; ws.cell(r, 4).value = row['승인일자']
        ws.cell(r, 5).value = row['사업자등록번호']; ws.cell(r, 6).value = row['거래처명']
        ws.cell(r, 7).value = row['거래처유형'] or None
        ws.cell(r, 8).value = int(row['공급가액']); ws.cell(r, 9).value = int(row['세액'])
        ws.cell(r, 10).value = int(row['봉사료']); ws.cell(r, 11).value = int(row['합계금액'])
        ws.cell(r, 12).value = row['부가세공제여부']
        ws.cell(r, 13).value = int(row['부가세유형']) if row['부가세유형'] != "" else None
        ws.cell(r, 14).value = row['계정과목']
        ws.cell(r, 15).value = row['품목(적요)'] or None
    ws.cell(9, 8).value = int(rows_df['공급가액'].sum())
    ws.cell(9, 9).value = int(rows_df['세액'].sum())
    ws.cell(9, 10).value = int(rows_df['봉사료'].sum())
    ws.cell(9, 11).value = int(rows_df['합계금액'].sum())
    wb.save(tmp_path); fix_html_entities(tmp_path)
    with open(tmp_path, 'rb') as f: data = f.read()
    os.remove(tmp_path)
    return data


# ── 탭 구성 ──────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📊 경비율 계산기", "🧾 종합소득세 계산기", "💳 카드사 변환기"])


# ════════════════════════════════════════════════════
# TAB 1 : 경비율 계산기
# ════════════════════════════════════════════════════
with tab1:
    st.title("📊 단순경비율 계정과목 배분 계산기")
    st.caption("업종코드와 매출액을 입력하면 Claude AI가 계정과목별 경비를 산정합니다.")

    with st.form("calc_form"):
        col1, col2, col3 = st.columns(3)
        with col1:
            업종코드 = st.text_input("업종코드", placeholder="예) 552101")
        with col2:
            매출액_input = st.text_input("매출액 (원)", placeholder="예) 150,000,000")
        with col3:
            경비율_input = st.text_input("경비율 (%)", placeholder="미입력 시 단순경비율 자동적용",
                                         help="직접 입력하면 해당 경비율로 계정과목 배분. 비워두면 단순경비율 자동 사용.")

        submitted = st.form_submit_button("계산하기", use_container_width=True, type="primary")

    if submitted:
        if not 업종코드 or not 매출액_input:
            st.warning("업종코드와 매출액을 모두 입력해주세요.")
        else:
            try:
                매출액 = int(매출액_input.replace(",", "").replace(" ", ""))
            except ValueError:
                st.error("매출액은 숫자로 입력해주세요.")
                st.stop()

            # 경비율 처리 (입력 없으면 None → 단순경비율 사용)
            커스텀경비율 = None
            if 경비율_input.strip():
                try:
                    커스텀경비율 = float(경비율_input.replace("%", "").strip())
                    if not (0 < 커스텀경비율 <= 100):
                        st.error("경비율은 0~100 사이 값을 입력해주세요.")
                        st.stop()
                except ValueError:
                    st.error("경비율은 숫자로 입력해주세요. 예) 64.1")
                    st.stop()

            df = load_data()
            accounts = load_accounts()
            business = find_business(df, 업종코드.strip())

            if business is None:
                st.error(f"업종코드 **{업종코드}** 를 찾을 수 없습니다.")
                st.stop()

            use_purchase = has_product_purchase(business["중분류"])
            적용경비율 = 커스텀경비율 if 커스텀경비율 else float(business["단순일반율"])

            st.divider()
            st.subheader("📌 업종 정보")
            c1, c2, c3 = st.columns(3)
            c1.metric("업종명", business["세세분류"])
            c2.metric("단순경비율(일반)", f"{business['단순일반율']}%")
            c3.metric("적용 경비율", f"{적용경비율}%",
                      delta="직접 입력" if 커스텀경비율 else "단순경비율 자동적용",
                      delta_color="normal" if 커스텀경비율 else "off")
            st.caption(f"중분류: {business['중분류']} / 소분류: {business['소분류']} / 세분류: {business['세분류']}")

            st.divider()
            with st.spinner("Claude AI가 계정과목별 비율을 산정하고 있습니다..."):
                try:
                    distribution = get_expense_distribution(business, accounts, use_purchase, 적용경비율)
                except Exception as e:
                    st.error(f"API 오류: {e}")
                    st.stop()

            st.subheader("📋 계정과목별 경비 내역")
            rows = []
            total_rate = 0
            total_amount = 0
            for account, rate in distribution.items():
                if rate > 0:
                    amount = 매출액 * rate / 100
                    rows.append({"계정과목": account, "비율": f"{rate:.1f}%", "금액 (원)": f"{amount:,.0f}"})
                    total_rate += rate
                    total_amount += amount

            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            st.divider()
            c1, c2, c3 = st.columns(3)
            c1.metric("매출액", f"{매출액:,.0f}원")
            c2.metric("총 경비", f"{total_amount:,.0f}원", f"{total_rate:.1f}%")
            소득금액 = 매출액 - total_amount
            c3.metric("소득금액 (매출 - 경비)", f"{소득금액:,.0f}원", f"{100 - total_rate:.1f}%")
            st.caption(f"※ 단순경비율 {business['단순일반율']}% 기준 / Claude AI 추정값으로 실제 신고와 다를 수 있습니다.")

            # 종합소득세 탭으로 넘길 소득금액 저장
            st.session_state["소득금액"] = 소득금액
            st.info("💡 종합소득세를 계산하려면 상단 **'🧾 종합소득세 계산기'** 탭을 클릭하세요.")


# ════════════════════════════════════════════════════
# TAB 2 : 종합소득세 계산기
# ════════════════════════════════════════════════════
with tab2:
    st.title("🧾 종합소득세 계산기")
    st.caption("소득금액과 공제항목을 입력하면 종합소득세를 자동 계산합니다.")

    # 소득금액 (경비율 탭에서 넘어온 값 or 직접 입력)
    default_소득 = st.session_state.get("소득금액", 0)

    st.subheader("① 소득금액")
    소득금액_input = st.number_input(
        "소득금액 (원)",
        min_value=0,
        value=int(default_소득),
        step=100000,
        format="%d",
        help="경비율 계산기에서 자동으로 불러오거나 직접 입력하세요."
    )

    st.divider()

    # ── 소득공제 ──
    st.subheader("② 소득공제")
    dc1, dc2, dc3 = st.columns(3)
    with dc1:
        국민연금 = st.number_input("국민연금보험료 (원)", min_value=0, value=0, step=10000, format="%d")
    with dc2:
        소상공인공제 = st.number_input("소상공인공제부금 (원)", min_value=0, value=0, step=10000, format="%d",
                                       help="노란우산공제 납입액")
    with dc3:
        인적공제 = st.number_input("인적공제 (원)", min_value=0, value=1500000, step=500000, format="%d",
                                    help="본인 150만원 기본, 부양가족 1인당 150만원 추가")

    총소득공제 = 국민연금 + 소상공인공제 + 인적공제

    st.divider()

    # ── 세액공제 ──
    st.subheader("③ 세액공제")
    tc1, tc2, tc3 = st.columns(3)
    with tc1:
        퇴직연금공제 = st.number_input("퇴직연금세액공제 (원)", min_value=0, value=0, step=10000, format="%d")
    with tc2:
        연금계좌공제 = st.number_input("연금계좌세액공제 (원)", min_value=0, value=0, step=10000, format="%d")
    with tc3:
        표준세액공제 = 70000
        st.text_input("표준세액공제 (원)", value="70,000", disabled=True, help="고정값 70,000원")

    총세액공제 = 퇴직연금공제 + 연금계좌공제 + 표준세액공제

    st.divider()

    # ── 기납부세액 ──
    st.subheader("④ 기납부세액")
    pc1, pc2 = st.columns(2)
    with pc1:
        중간예납 = st.number_input("중간예납세액 (원)", min_value=0, value=0, step=10000, format="%d")
    with pc2:
        원천징수 = st.number_input("원천징수세액 (원)", min_value=0, value=0, step=10000, format="%d")

    총기납부 = 중간예납 + 원천징수

    st.divider()

    # ── 계산 버튼 ──
    if st.button("종합소득세 계산", use_container_width=True, type="primary"):

        과세표준 = max(소득금액_input - 총소득공제, 0)
        산출세액 = calc_income_tax(과세표준)
        납부세액 = max(산출세액 - 총세액공제, 0)
        최종납부 = max(납부세액 - 총기납부, 0)
        환급세액 = max(총기납부 - 납부세액, 0)
        지방소득세 = round(납부세액 * 0.1)

        # 해당 세율 구간 찾기
        for limit, rate, deduction in TAX_BRACKETS:
            if 과세표준 <= limit:
                적용세율 = rate
                누진공제액 = deduction
                break

        st.divider()
        st.subheader("📊 종합소득세 계산 결과")

        # 계산 구조 표
        steps = [
            ("소득금액",           f"{소득금액_input:>20,.0f} 원", ""),
            ("(-) 소득공제",       f"{총소득공제:>20,.0f} 원",
             f"국민연금 {국민연금:,.0f} + 소상공인 {소상공인공제:,.0f} + 인적공제 {인적공제:,.0f}"),
            ("= 과세표준",         f"{과세표준:>20,.0f} 원", ""),
            (f"× 세율 ({적용세율*100:.0f}%)", f"{round(과세표준 * 적용세율):>20,.0f} 원", ""),
            (f"(-) 누진공제",      f"{누진공제액:>20,.0f} 원", ""),
            ("= 산출세액",         f"{산출세액:>20,.0f} 원", ""),
            ("(-) 세액공제",       f"{총세액공제:>20,.0f} 원",
             f"퇴직연금 {퇴직연금공제:,.0f} + 연금계좌 {연금계좌공제:,.0f} + 표준 {표준세액공제:,.0f}"),
            ("= 납부세액",         f"{납부세액:>20,.0f} 원", ""),
            ("(-) 기납부세액",     f"{총기납부:>20,.0f} 원",
             f"중간예납 {중간예납:,.0f} + 원천징수 {원천징수:,.0f}"),
        ]

        for label, value, note in steps:
            r1, r2, r3 = st.columns([3, 3, 4])
            r1.write(f"**{label}**")
            r2.write(value)
            if note:
                r3.caption(note)

        st.divider()

        # 최종 결과 카드
        res1, res2, res3 = st.columns(3)
        if 환급세액 > 0:
            res1.metric("🔴 종합소득세 납부세액", f"{최종납부:,.0f}원")
            res2.metric("💚 환급세액", f"{환급세액:,.0f}원")
        else:
            res1.metric("🔴 종합소득세 납부세액", f"{최종납부:,.0f}원")
            res2.metric("🏙️ 지방소득세 (10%)", f"{지방소득세:,.0f}원")

        res3.metric("💳 총 세금 부담", f"{최종납부 + 지방소득세:,.0f}원",
                    help="종합소득세 + 지방소득세 합계")

        st.caption(f"※ 적용 세율: {적용세율*100:.0f}% / 누진공제: {누진공제액:,.0f}원 / 지방소득세는 납부세액의 10%")
        st.caption("※ 본 계산은 추정값이며 실제 세액은 세무사 확인 후 신고하시기 바랍니다.")


# ════════════════════════════════════════════════════
# TAB 3 : 카드사 변환기
# ════════════════════════════════════════════════════
with tab3:
    st.title("💳 카드사 통합 변환기")
    st.caption("카드사 엑셀 파일을 업로드하면 세무사랑 업로드용 파일로 자동 변환합니다.")

    st.info("📌 파일명 형식: **상호명_사업자번호_카드사_카드번호_직원유무_차량유무.xlsx**\n\n예) 용은물류_3020895715_삼성카드_5120-2800-0000-5697_직원없음_차량있음.xlsx")

    st.subheader("① 카드사 파일 업로드")
    st.caption("삼성/하나/신한/비씨/NH농협/현대/국민/IBK기업BC/카카오뱅크/롯데카드 지원 / 여러 파일 동시 업로드 가능")
    uploaded_cards = st.file_uploader(
        "카드사 엑셀 파일 선택",
        type=["xlsx"],
        accept_multiple_files=True,
        key="card_files"
    )

    if uploaded_cards:
        st.divider()
        if st.button("🔄 변환 시작", use_container_width=True, type="primary"):
            all_rows, all_stats = [], []
            errors = []

            with st.spinner("변환 중..."):
                for uf in uploaded_cards:
                    info = parse_filename_card(uf.name)
                    card_co = info["신용카드사명"]
                    card_no = info["신용카드번호"]
                    file_bytes = uf.read()

                    try:
                        if "삼성" in card_co:
                            rows, stats = parse_samsung_card(file_bytes, card_co, card_no)
                        elif "하나" in card_co:
                            rows, stats = parse_hana_card(file_bytes, card_co, card_no)
                        elif "신한" in card_co:
                            rows, stats = parse_shinhan_card(file_bytes, card_co, card_no)
                        elif "비씨" in card_co or "BC" in card_co.upper():
                            rows, stats = parse_bc_card(file_bytes, card_co, card_no)
                        elif "농협" in card_co or "NH" in card_co.upper():
                            rows, stats = parse_nh_card(file_bytes, card_co, card_no)
                        elif "현대" in card_co:
                            rows, stats = parse_hyundai_card(file_bytes, card_co, card_no)
                        elif "국민" in card_co or "KB" in card_co.upper():
                            rows, stats = parse_kb_card(file_bytes, card_co, card_no)
                        elif "기업" in card_co or "IBK" in card_co.upper():
                            rows, stats = parse_ibk_bc_card(file_bytes, card_co, card_no)
                        elif "카카오" in card_co or "카카오" in uf.name:
                            rows, stats = parse_kakao_card(file_bytes, card_co or "카카오뱅크", card_no)
                        elif "롯데" in card_co or "롯데" in uf.name:
                            rows, stats = parse_lotte_card(file_bytes, card_co or "롯데카드", card_no)
                        else:
                            # 카드사 미입력 시 파일 내용으로 자동 감지
                            try:
                                df_auto = pd.read_excel(io.BytesIO(file_bytes), header=None, dtype=str)
                                matched = False
                                for _i, _row in df_auto.iterrows():
                                    _vals = [str(v).strip() for v in _row if pd.notna(v) and str(v).strip() not in ['', 'nan']]
                                    if '접수일자' in _vals and '가맹점명' in _vals and '이용금액' in _vals:
                                        rows, stats = parse_ibk_bc_card(file_bytes, "IBK기업BC카드", card_no)
                                        matched = True
                                        break
                                    if '거래일시' in _vals and '가맹점명' in _vals and '취소여부' in _vals:
                                        rows, stats = parse_kakao_card(file_bytes, "카카오뱅크", card_no)
                                        matched = True
                                        break
                                    if '매출일자' in _vals and '가맹점명' in _vals and '가맹점번호' in _vals:
                                        rows, stats = parse_lotte_card(file_bytes, "롯데카드", card_no)
                                        matched = True
                                        break
                                if not matched:
                                    errors.append(f"⚠️ {uf.name}: 지원하지 않는 카드사 ({card_co})\n지원: 삼성/하나/신한/비씨/NH농협/현대/국민/IBK기업BC/카카오뱅크/롯데카드")
                                    continue
                            except Exception:
                                errors.append(f"⚠️ {uf.name}: 지원하지 않는 카드사 ({card_co})\n지원: 삼성/하나/신한/비씨/NH농협/현대/국민/IBK기업BC/카카오뱅크/롯데카드")
                                continue
                        all_rows.append(rows)
                        all_stats.append(stats)
                    except Exception as e:
                        errors.append(f"❌ {uf.name}: {e}")

            if errors:
                for e in errors: st.error(e)

            non_empty_rows  = [r for r in all_rows  if isinstance(r, pd.DataFrame) and not r.empty]
            non_empty_stats = [s for s in all_stats if isinstance(s, pd.DataFrame) and not s.empty]
            if not non_empty_rows and not errors:
                st.warning("⚠️ 변환할 거래 내역이 없습니다. 파일 형식을 확인해주세요.")
            if non_empty_rows:
                combined = pd.concat(non_empty_rows, ignore_index=True)
                stats_df = pd.concat(non_empty_stats, ignore_index=True)

                # ── AI 자동 분류 (미분류 항목) ──
                ai_classified_cnt = 0
                unclassified_mask = stats_df['계정과목'] == '미분류'
                if unclassified_mask.any():
                    with st.spinner("🤖 미분류 항목 AI 분류 중..."):
                        try:
                            unc_vendors = stats_df.loc[unclassified_mask, '가맹점명'].tolist()
                            ai_map = classify_unclassified_with_claude(unc_vendors)
                            if ai_map:
                                for idx in stats_df[unclassified_mask].index:
                                    vendor = stats_df.loc[idx, '가맹점명']
                                    new_acc = ai_map.get(vendor)
                                    if new_acc and new_acc in ACCOUNT_CODE_MAP:
                                        stats_df.loc[idx, '계정과목'] = new_acc
                                        stats_df.loc[idx, '분류방법'] = 'AI'
                                        combined.loc[idx, '계정과목'] = ACCOUNT_CODE_MAP[new_acc]
                                        vat = "불공제" if new_acc == "접대비" else "공제"
                                        combined.loc[idx, '부가세공제여부'] = vat
                                        combined.loc[idx, '부가세유형'] = 57 if vat == "공제" else ""
                                        ai_classified_cnt += 1
                        except Exception as ai_err:
                            st.warning(f"⚠️ AI 분류 중 오류 발생: {ai_err}")

                total_cnt = len(combined)
                classified = len(stats_df[stats_df['계정과목'] != '미분류']) if '계정과목' in stats_df.columns else 0
                unclassified = total_cnt - classified
                first_info = parse_filename_card(uploaded_cards[0].name)

                st.divider()
                st.subheader("📊 변환 결과")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("총 거래건수", f"{total_cnt:,}건")
                c2.metric("키워드 분류", f"{classified - ai_classified_cnt:,}건")
                c3.metric("🤖 AI 분류", f"{ai_classified_cnt:,}건")
                c4.metric("미분류", f"{unclassified:,}건", delta_color="inverse")

                amt_col = st.columns(3)
                amt_col[0].metric("공급가액 합계", f"{combined['공급가액'].sum():,.0f}원")
                amt_col[1].metric("세액 합계", f"{combined['세액'].sum():,.0f}원")
                amt_col[2].metric("합계금액", f"{combined['합계금액'].sum():,.0f}원")

                # 계정과목별 집계
                st.divider()
                st.subheader("📋 계정과목별 집계")
                summary = stats_df.groupby('계정과목')['금액'].agg(['count','sum']).reset_index()
                summary.columns = ['계정과목', '건수', '금액합계']
                summary['금액합계'] = summary['금액합계'].apply(lambda x: f"{x:,.0f}원")
                st.dataframe(summary, use_container_width=True, hide_index=True)

                if unclassified > 0:
                    st.warning(f"⚠️ 미분류 {unclassified}건 — 리포트 파일에서 확인 후 수동 수정하세요.")
                elif ai_classified_cnt > 0:
                    st.success(f"✅ AI가 미분류 {ai_classified_cnt}건을 자동 분류했습니다. 리포트에서 검토를 권장합니다.")

                st.divider()
                st.subheader("⬇️ 파일 다운로드")

                # 세무사랑 파일 생성
                try:
                    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
                    세무사랑_data = write_to_template(
                        combined,
                        first_info["업체명"],
                        first_info["사업자번호"]
                    )
                    dl1, dl2 = st.columns(2)
                    with dl1:
                        st.download_button(
                            label="📥 세무사랑 업로드 파일",
                            data=세무사랑_data,
                            file_name=f"세무사랑_통합_{ts}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
                    with dl2:
                        buf = io.BytesIO()
                        stats_df.to_excel(buf, index=False, engine='openpyxl')
                        st.download_button(
                            label="📊 분류 리포트",
                            data=buf.getvalue(),
                            file_name=f"분류리포트_{ts}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                except Exception as e:
                    st.error(f"파일 생성 오류: {e}")
                    st.info("세무사랑 템플릿 없이 리포트만 다운로드 가능합니다.")
                    buf = io.BytesIO()
                    stats_df.to_excel(buf, index=False, engine='openpyxl')
                    st.download_button("📊 분류 리포트 다운로드", buf.getvalue(), f"분류리포트_{ts}.xlsx")
